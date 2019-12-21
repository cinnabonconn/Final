# This is my final project for ITEC 1150. 
# I was not able to use JSON with this, but i think my solution was a good way to go about it.
# I also was able to get the second extra credit part to work, which was exciting.

# You need imports for the program to work
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

# Variables
dictionary = {} #An empty dictionary
id = []
course_number = []
title = []
day = []
time = []
credits = []
instructor = []
location = []

# Gathering API information
url = 'https://eservices.minnstate.edu/registration/search/advancedSubmit.html?campusid=305&' \
      'searchrcid=0305&searchcampusid=305&yrtr=20205&subject=ITEC&courseNumber=&courseId=&op' \
      'enValue=OPEN_PLUS_WAITLIST&delivery=ALL&showAdvanced=&starttime=&endtime=&mntransfer=' \
      '&credittype=ALL&credits=&instructor=&keyword=&begindate=&site=&resultNumber=250'

r = requests.get(url)  # Requesting the URL HTML information
soup = BeautifulSoup(r.content, 'html.parser')  # Reading HTML and making sense of its structure
container = soup.find(attrs={'id':'searchResultsContainer'})
results = soup.find_all("div", {"class" : "meta"})  # Making a list of all 37 classes data
results2 = soup.find_all('div', attrs={'style':'white-space: nowrap;overflow: hidden;'})  # Making a list of all 37 classes data
results3 = soup.find_all('td')

# ID DONE
i = 0  # Variable for iterating
for data in results:  # Running through every result
      data = results[i].get('id')  # Print the class out
      id.append(data[-6:])  # Adding all the ids to an array
      i = i + 1  # Makes the list iterate every class

# COURSE NUMBER
i = 10  # Variable for iterating
z = 11
for data in results:  # Running through every result
      data = results3[i]  # Print the class out
      data2 = results3[z]
      course_number.append(str(data)[6:10] + '-' +str(data2)[6:8])  # Adding all the courses to an array
      i = i + 14  # Makes the list iterate every class
      z = z + 14  # Makes the list iterate every class

# TITLE DONE
i = 0  # Variable for iterating
for data in results:  # Running through every result
      data = results[i].get('id')  # Print the class out
      title.append(data[:-6])  # Adding all the titles to an array
      i = i + 1  # Makes the list iterate every class

# DAY
i = 14  # Variable for iterating
for data in results:  # Running through every result
      data = results3[i]  # Print the credits out
      if "Monday" in str(data):
            day.append(str(data)[70:76])  # Adding all the days to an array
      elif "Tuesday" in str(data):
            day.append(str(data)[70:77])  # Adding all the days to an array
      elif "Wednesday" in str(data):
            day.append(str(data)[70:79])  # Adding all the days to an array
      elif "Thursday" in str(data):
            day.append(str(data)[70:78])  # Adding all the days to an array
      elif "Friday" in str(data):
            day.append(str(data)[70:76])  # Adding all the days to an array
      elif "Saturday" in str(data):
            day.append(str(data)[70:78])  # Adding all the days to an array
      elif "Sunday" in str(data):
            day.append(str(data)[70:76])  # Adding all the days to an array
      elif "not available" in str(data):
            day.append(str(data)[70:83])  # Adding all the days to an array
      i = i + 14  # Makes the list iterate every class

# TIME
i = 15  # Variable for iterating
for data in results:  # Running through every result
      data = results3[i]  # Print the credits out
      if "9:40pm" in str(data):
            time.append(str(data)[92:98] + '-' + str(data)[101:107])  # Adding all the times to an array
      elif "4:10pm" in str(data):
            time.append(str(data)[92:99] + '-' + str(data)[102:108])  # Adding all the times to an array
      elif "6:40pm" in str(data):
            time.append(str(data)[92:98] + '-' + str(data)[101:107])  # Adding all the times to an array
      elif "9:35pm" in str(data):
            time.append(str(data)[92:98] + '-' + str(data)[101:107])  # Adding all the times to an array
      elif "3:00pm" in str(data):
            time.append(str(data)[92:99] + '-' + str(data)[102:108])  # Adding all the times to an array
      elif "2:05pm" in str(data):
            time.append(str(data)[92:98] + '-' + str(data)[102:108])  # Adding all the times to an array
      elif "1:30pm" in str(data):
            time.append(str(data)[92:98] + '-' + str(data)[101:107])  # Adding all the times to an array
      elif "1:10pm" in str(data):
            time.append(str(data)[92:98] + '-' + str(data)[102:107])  # Adding all the times to an array
      elif "4:50pm" in str(data):
            time.append(str(data)[92:98] + '-' + str(data)[102:108])  # Adding all the times to an array
      elif "11:05am" in str(data):
            time.append(str(data)[92:98] + '-' + str(data)[101:108])  # Adding all the times to an array
      elif "9:00pm" in str(data):
            time.append(str(data)[92:98] + '-' + str(data)[101:107])  # Adding all the times to an array
      elif "11:59am" in str(data):
            time.append(str(data)[92:98] + '-' + str(data)[101:108])  # Adding all the times to an array
      elif "Arranged" in str(data):
            time.append(str(data)[75:83])  # Adding all the instructors to an array
      i = i + 14  # Makes the list iterate every class

#CREDITS
i = 16  # Variable for iterating
for data in results:  # Running through every result
      data = results3[i]  # Print the credits out
      credits.append(str(data)[36:39])  # Adding all the courses to an array
      i = i + 14  # Makes the list iterate every class

# INSTRUCTOR
i = 18  # Variable for iterating
for data in results:  # Running through every result
      data = results3[i]  # Print the credits out
      if "Brian" in str(data):
            instructor.append(str(data)[54:62] + ' ' + str(data)[63:68])  # Adding all the instructors to an array
      elif "Christoph" in str(data):
            instructor.append(str(data)[54:64] + ' ' + str(data)[65:74])  # Adding all the instructors to an array
      elif "Catherine" in str(data):
            instructor.append(str(data)[54:59] + ' ' + str(data)[60:69])  # Adding all the instructors to an array
      elif "Mary" in str(data):
            instructor.append(str(data)[54:59] + ' ' + str(data)[60:64])  # Adding all the instructors to an array
      elif "Steven" in str(data):
            instructor.append(str(data)[54:60] + ' ' + str(data)[61:67])  # Adding all the instructors to an array
      elif "Duncan" in str(data):
            instructor.append(str(data)[54:58] + ' ' + str(data)[59:65])  # Adding all the instructors to an array
      elif "Clara" in str(data):
            instructor.append(str(data)[54:60] + ' ' + str(data)[61:66])  # Adding all the instructors to an array
      elif "Warren" in str(data):
            instructor.append(str(data)[54:63] + ' ' + str(data)[64:70])  # Adding all the instructors to an array
      elif "Justin" in str(data):
            instructor.append(str(data)[54:62] + ' ' + str(data)[63:69])  # Adding all the instructors to an array
      elif "Staff" in str(data):
            instructor.append(str(data)[54:60] + ' ' + str(data)[61:66])  # Adding all the instructors to an array
      i = i + 14  # Makes the list iterate every class

# LOCATION
i = 20  # Variable for iterating
for data in results:  # Running through every result
      data = results3[i]  # Print the credits out
      if "Online" in str(data):
            location.append("Online")  # Adding all the locations to an array
      else:
            location.append("MCTC " + str(data)[141:156])  # Adding all the locations to an array
      i = i + 14  # Makes the list iterate every class

print(location)

# Make a new workbook
workbook = Workbook()

# Get the active sheet
worksheet = workbook.active

worksheet.title = 'ITEC Courses Information'

# Adding titles to all the excel columns
worksheet.cell(1, 1, "ID #")
worksheet.cell(1, 2, "Course ID")
worksheet.cell(1, 3, "Title")
worksheet.cell(1, 4, "Day")
worksheet.cell(1, 5, "Time")
worksheet.cell(1, 6, "Credits")
worksheet.cell(1, 7, "Instructor")
worksheet.cell(1, 8, "Location")

# Assigning the data to each row
for index, data in enumerate(id):  # Iterating through the list
    worksheet.cell(index + 2, 1, data)  # Assigning to certain spots on the Excel Spreadsheet

for index, data in enumerate(course_number):  # Iterating through the list
    worksheet.cell(index + 2, 2, data)  # Assigning to certain spots on the Excel Spreadsheet

for index, data in enumerate(title):  # Iterating through the list
    worksheet.cell(index + 2, 3, data)  # Assigning to certain spots on the Excel Spreadsheet

for index, data in enumerate(day):  # Iterating through the list
    worksheet.cell(index + 2, 4, data)  # Assigning to certain spots on the Excel Spreadsheet

for index, data in enumerate(time):  # Iterating through the list
    worksheet.cell(index + 2, 5, data)  # Assigning to certain spots on the Excel Spreadsheet

for index, data in enumerate(credits):  # Iterating through the list
    worksheet.cell(index + 2, 6, data)  # Assigning to certain spots on the Excel Spreadsheet

for index, data in enumerate(instructor):  # Iterating through the list
    worksheet.cell(index + 2, 7, data)  # Assigning to certain spots on the Excel Spreadsheet

for index, data in enumerate(location):  # Iterating through the list
    worksheet.cell(index + 2, 8, data)  # Assigning to certain spots on the Excel Spreadsheet

# Save Workbook
workbook.save('ITEC_Courses.xlsx')
